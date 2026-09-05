# -*- coding: utf-8 -*-
"""Builds the tester-facing workbook from the same case data as the PDF.

Run:  python build_qa_xlsx.py        (writes ../desk-portal-test-cases.xlsx)
Needs: pip install openpyxl

Testers fill in Result, Tester, Date and Notes on the Test Cases sheet. Everything else is
reference material, and the Summary sheet is entirely formulas so it keeps itself current.
"""
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import qa_plan_part1  # noqa: F401  registers modules 1-6
import qa_plan_part2  # noqa: F401
import qa_plan_part3  # noqa: F401
import qa_plan_part4  # noqa: F401
from qa_plan_part1 import MODULES
from qa_plan_sql import ACCESS, LOG_LINES, SQL_COOKBOOK

FONT = "Arial"
INK = "14532D"
BG_HEAD = "14532D"
BG_BAND = "F4F7F4"
EDIT_FILL = PatternFill("solid", fgColor="FFF9DB")   # cells a tester fills in
PASS_FILL = PatternFill("solid", fgColor="D7F0DC")
FAIL_FILL = PatternFill("solid", fgColor="FBD9D9")
BLOCK_FILL = PatternFill("solid", fgColor="FDECD2")
THIN = Side(style="thin", color="CBD8CE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RESULTS = ["Not run", "Pass", "Fail", "Blocked", "N/A"]

wb = Workbook()

# --------------------------------------------------------------- Test Cases
ws = wb.active
ws.title = "Test Cases"

HEADERS = ["Module", "ID", "Test", "How to test", "Expected result", "Verify with",
           "Result", "Tester", "Date", "Notes / defect ref"]
WIDTHS = [30, 12, 34, 62, 52, 46, 12, 16, 12, 34]

for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=BG_HEAD)
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    c.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26

row = 2
for mod_name, _intro, cases in MODULES:
    for cid, title, steps, expected, verify in cases:
        values = [mod_name, cid, title, steps, expected, verify, "Not run", "", "", ""]
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name=FONT, size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if i in (7, 8, 9, 10):          # the columns a tester edits
                c.fill = EDIT_FILL
        ws.cell(row=row, column=2).font = Font(name=FONT, size=9, bold=True, color="EA580C")
        ws.cell(row=row, column=9).number_format = "yyyy-mm-dd"
        row += 1

last = row - 1

dv = DataValidation(type="list", formula1='"%s"' % ",".join(RESULTS), allow_blank=True)
dv.prompt = "Pick a result"
dv.promptTitle = "Result"
ws.add_data_validation(dv)
dv.add("G2:G%d" % last)

for op, fill in (("Pass", PASS_FILL), ("Fail", FAIL_FILL), ("Blocked", BLOCK_FILL)):
    ws.conditional_formatting.add(
        "G2:G%d" % last,
        CellIsRule(operator="equal", formula=['"%s"' % op], fill=fill))

ws.freeze_panes = "C2"
ws.auto_filter.ref = "A1:J%d" % last

# --------------------------------------------------------------- Summary
sm = wb.create_sheet("Summary")
sm.column_dimensions["A"].width = 46
for col in "BCDEFG":
    sm.column_dimensions[col].width = 12

sm["A1"] = "Progress by module"
sm["A1"].font = Font(name=FONT, bold=True, size=13, color=INK)
sm["A2"] = ("Every figure here is a formula over the Test Cases sheet, so it updates itself as "
            "testers fill in the Result column. Nothing on this sheet should be typed into.")
sm["A2"].font = Font(name=FONT, size=9, italic=True, color="5B6B60")
sm.merge_cells("A2:G2")
sm["A2"].alignment = Alignment(wrap_text=True, vertical="top")
sm.row_dimensions[2].height = 26

head = ["Module", "Cases", "Pass", "Fail", "Blocked", "Not run", "% done"]
for i, h in enumerate(head, start=1):
    c = sm.cell(row=4, column=i, value=h)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill("solid", fgColor=BG_HEAD)
    c.border = BORDER
    c.alignment = Alignment(horizontal="left" if i == 1 else "center")

r = 5
CASES = "'Test Cases'!$A$2:$A$%d" % last
RES = "'Test Cases'!$G$2:$G$%d" % last
for mod_name, _intro, cases in MODULES:
    sm.cell(row=r, column=1, value=mod_name).font = Font(name=FONT, size=9)
    sm.cell(row=r, column=2, value='=COUNTIF(%s,$A%d)' % (CASES, r))
    for i, res in enumerate(["Pass", "Fail", "Blocked", "Not run"], start=3):
        sm.cell(row=r, column=i,
                value='=COUNTIFS(%s,$A%d,%s,"%s")' % (CASES, r, RES, res))
    # Guard the denominator: a module with no rows would divide by zero.
    sm.cell(row=r, column=7,
            value='=IFERROR(($C%d+$D%d+$E%d)/$B%d,0)' % (r, r, r, r))
    sm.cell(row=r, column=7).number_format = "0%"
    for i in range(1, 8):
        cell = sm.cell(row=r, column=i)
        cell.border = BORDER
        if i > 1:
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(horizontal="center")
        if r % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=BG_BAND)
    r += 1

sm.cell(row=r, column=1, value="Total").font = Font(name=FONT, bold=True, size=10, color=INK)
for i in range(2, 7):
    col = get_column_letter(i)
    sm.cell(row=r, column=i, value="=SUM(%s5:%s%d)" % (col, col, r - 1))
sm.cell(row=r, column=7, value='=IFERROR(($C%d+$D%d+$E%d)/$B%d,0)' % (r, r, r, r))
sm.cell(row=r, column=7).number_format = "0%"
for i in range(1, 8):
    c = sm.cell(row=r, column=i)
    c.font = Font(name=FONT, bold=True, size=10)
    c.border = Border(left=THIN, right=THIN, top=Side(style="medium", color=INK), bottom=THIN)
    if i > 1:
        c.alignment = Alignment(horizontal="center")
sm.freeze_panes = "A5"

# --------------------------------------------------------------- How to use
hu = wb.create_sheet("How to use")
hu.column_dimensions["A"].width = 22
hu.column_dimensions["B"].width = 110


def block(sheet, r, label, text, bold=False):
    a = sheet.cell(row=r, column=1, value=label)
    a.font = Font(name=FONT, bold=True, size=10, color=INK)
    a.alignment = Alignment(vertical="top")
    b = sheet.cell(row=r, column=2, value=text)
    b.font = Font(name=FONT, size=10, bold=bold)
    b.alignment = Alignment(wrap_text=True, vertical="top")
    return r + 1


hu["A1"] = "Desk Portal / PIO Manage - test case workbook"
hu["A1"].font = Font(name=FONT, bold=True, size=14, color=INK)
hu.merge_cells("A1:B1")

r = 3
r = block(hu, r, "Fill in", "On the Test Cases sheet, only the four SHADED columns are yours: "
                            "Result, Tester, Date and Notes / defect ref. Everything to their "
                            "left is the plan and should not be edited here - change it in the "
                            "generator so the PDF and this workbook stay in step.")
r = block(hu, r, "Result", "Pick from the dropdown: Not run, Pass, Fail, Blocked, N/A. The cell "
                           "colours itself, and the Summary sheet recounts automatically.")
r = block(hu, r, "Filtering", "Row 1 has filters. Filter Module to work one area at a time, or "
                              "Result = Fail to review defects.")
r += 1
r = block(hu, r, "Three rules", "These decide whether a pass means anything in this system:", True)
for t in [
    "1. Confirm the build under test is newer than the fix you are testing. More time has been "
    "lost here to stale builds than to real defects.",
    "2. A check that matches something already true is not a check. When waiting on a sync or a "
    "rollup, capture the current timestamp first and wait for a DIFFERENT one.",
    "3. Absence of an error is not evidence. An unmapped value, a skipped ticket and a truncated "
    "import all look exactly like normal successful operation.",
]:
    r = block(hu, r, "", t)

r += 1
r = block(hu, r, "Example row", "How a completed row should look:", True)
ex_head = ["ID", "Result", "Tester", "Date", "Notes / defect ref"]
for i, h in enumerate(ex_head, start=1):
    c = hu.cell(row=r, column=i, value=h)
    c.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BG_HEAD)
    c.border = BORDER
r += 1
for i, v in enumerate(["SYNC-02", "Fail", "A. Tester", "2026-09-08",
                       "Imported 100 of 135. Raised DESK-412."], start=1):
    c = hu.cell(row=r, column=i, value=v)
    c.font = Font(name=FONT, size=9)
    c.border = BORDER
    c.fill = EDIT_FILL if i > 1 else PatternFill()
r += 2

r = block(hu, r, "Not covered", "Stated so nobody assumes otherwise: there is no SMTP anywhere "
                                "in the stack, so email delivery cannot be tested - enquiries "
                                "and reports are stored for in-portal download. Load testing, "
                                "DAST and penetration testing are authored but have never run "
                                "against the live stack. Disaster recovery is documented but "
                                "never rehearsed. ConnectWise closure dates cannot be verified "
                                "until a ticket is actually closed in that sandbox, and the "
                                "legal pages still contain unfilled placeholders.")

# --------------------------------------------------------------- Reference
ref = wb.create_sheet("Reference")
ref.column_dimensions["A"].width = 48
ref.column_dimensions["B"].width = 110

ref["A1"] = "Reference - access, SQL and log lines"
ref["A1"].font = Font(name=FONT, bold=True, size=13, color=INK)

r = 3
ref.cell(row=r, column=1, value="Getting in").font = Font(name=FONT, bold=True, size=11,
                                                          color=INK)
r += 1
for label, cmd in ACCESS:
    ref.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True, size=9)
    c = ref.cell(row=r, column=2, value=cmd)
    c.font = Font(name="Consolas", size=9)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
ref.cell(row=r, column=1, value="SQL cookbook").font = Font(name=FONT, bold=True, size=11,
                                                            color=INK)
r += 1
for label, q in SQL_COOKBOOK:
    a = ref.cell(row=r, column=1, value=label)
    a.font = Font(name=FONT, size=9, bold=True)
    a.alignment = Alignment(wrap_text=True, vertical="top")
    b = ref.cell(row=r, column=2, value=q)
    b.font = Font(name="Consolas", size=8.5)
    b.alignment = Alignment(wrap_text=True, vertical="top")
    for i in (1, 2):
        ref.cell(row=r, column=i).border = BORDER
    r += 1

r += 1
ref.cell(row=r, column=1, value="Log lines worth grepping").font = Font(name=FONT, bold=True,
                                                                        size=11, color=INK)
r += 1
for k, v in LOG_LINES:
    a = ref.cell(row=r, column=1, value=k)
    a.font = Font(name="Consolas", size=9)
    b = ref.cell(row=r, column=2, value=v)
    b.font = Font(name=FONT, size=9)
    b.alignment = Alignment(wrap_text=True, vertical="top")
    for i in (1, 2):
        ref.cell(row=r, column=i).border = BORDER
    r += 1

out_dir = os.environ.get("QA_OUT") or os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
out = os.path.join(out_dir, "desk-portal-test-cases.xlsx")
wb.save(out)
print("cases:", last - 1, "modules:", len(MODULES))
print("written:", out)
